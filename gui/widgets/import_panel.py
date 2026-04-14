# gui/widgets/import_panel.py
import os

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QComboBox, QFileDialog, QMessageBox, QGroupBox,
)
from PySide6.QtCore import Signal, Qt

from core.excel_importer import (
    import_from_excel, import_from_csv,
    detect_phone_column, detect_phone_column_csv,
    get_column_headers, deduplicate_numbers,
)
from gui.styles import COLORS


class DropZone(QLabel):
    """Label that accepts drag & drop of files."""

    file_dropped = Signal(str)

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setAcceptDrops(True)
        self.setText("Przeciągnij plik Excel/CSV tutaj lub użyj Ctrl+V")
        self.setAlignment(Qt.AlignCenter)
        self.setWordWrap(True)
        self.setMinimumHeight(32)
        self.setProperty("class", "dim")
        self._default_style = (
            f"border: 2px dashed {COLORS['border']}; "
            f"border-radius: 8px; "
            f"padding: 12px; "
            f"background-color: transparent; "
            f"color: {COLORS['text_secondary']};"
        )
        self._hover_style = (
            f"border: 2px dashed {COLORS['accent']}; "
            f"border-radius: 8px; "
            f"padding: 12px; "
            f"color: {COLORS['accent']}; "
            f"background-color: {COLORS['accent_light']};"
        )
        self.setStyleSheet(self._default_style)

    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            url = event.mimeData().urls()[0].toLocalFile()
            if url.lower().endswith((".xlsx", ".csv")):
                event.acceptProposedAction()
                self.setStyleSheet(self._hover_style)
                return
        event.ignore()

    def dragLeaveEvent(self, event):
        self.setStyleSheet(self._default_style)

    def dropEvent(self, event):
        self.setStyleSheet(self._default_style)
        url = event.mimeData().urls()[0].toLocalFile()
        self.file_dropped.emit(url)


class ImportPanel(QGroupBox):
    """Panel for importing phone numbers from Excel/CSV, drag&drop, clipboard."""

    numbers_changed = Signal(list, list, list)
    headers_changed = Signal(list)

    def __init__(self, settings=None, parent=None):
        super().__init__("Import numerów", parent)
        self._settings = settings
        self._current_path = ""
        self._headers = None
        self._build_ui()

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(12, 8, 12, 8)
        layout.setSpacing(6)

        row1 = QHBoxLayout()
        row1.setSpacing(12)

        self._btn_import = QPushButton("Wybierz plik")
        self._btn_import.setMinimumWidth(120)
        self._btn_import.clicked.connect(self._on_import_click)
        row1.addWidget(self._btn_import)

        lbl_col = QLabel("Kolumna:")
        lbl_col.setProperty("class", "dim")
        row1.addWidget(lbl_col)

        self._combo_column = QComboBox()
        self._combo_column.setMinimumWidth(150)
        self._combo_column.setEnabled(False)
        self._combo_column.currentIndexChanged.connect(self._on_column_changed)
        row1.addWidget(self._combo_column)

        self._lbl_file = QLabel("Nie wybrano pliku")
        self._lbl_file.setProperty("class", "dim")
        self._lbl_file.setStyleSheet("padding-left: 8px;")
        row1.addWidget(self._lbl_file)

        row1.addStretch()
        layout.addLayout(row1)

        self._drop_zone = DropZone()
        self._drop_zone.file_dropped.connect(self._load_file)
        layout.addWidget(self._drop_zone)

        self._lbl_summary = QLabel("")
        self._lbl_summary.setProperty("class", "dim")
        layout.addWidget(self._lbl_summary)

    def _on_import_click(self):
        initial_dir = ""
        if self._settings and self._settings.last_import_dir:
            initial_dir = self._settings.last_import_dir

        path, _ = QFileDialog.getOpenFileName(
            self, "Importuj plik",
            initial_dir,
            "Excel / CSV (*.xlsx *.csv);;Wszystkie pliki (*.*)",
        )
        if path:
            if self._settings:
                self._settings.last_import_dir = os.path.dirname(path)
            self._load_file(path)

    def _load_file(self, path: str):
        self._current_path = path
        filename = os.path.basename(path)
        is_csv = path.lower().endswith(".csv")

        try:
            if is_csv:
                detected = detect_phone_column_csv(path)
            else:
                detected = detect_phone_column(path)
                self._headers = get_column_headers(path)
        except Exception as e:
            QMessageBox.critical(self, "Błąd", str(e))
            return

        self._combo_column.blockSignals(True)
        self._combo_column.clear()

        max_cols = self._get_max_columns(path, is_csv)
        for i in range(max_cols):
            label = chr(65 + i) if i < 26 else f"Col{i + 1}"
            if self._headers and i < len(self._headers) and self._headers[i]:
                label = f"{chr(65 + i)} ({self._headers[i]})"
            if detected is not None and i == detected:
                label += " [auto]"
            self._combo_column.addItem(label, i)

        if detected is not None:
            self._combo_column.setCurrentIndex(detected)
        self._combo_column.setEnabled(True)
        self._combo_column.blockSignals(False)

        self._lbl_file.setText(filename)
        self._import_with_column(path, detected if detected is not None else 0)

    def _on_column_changed(self, index: int):
        if self._current_path and index >= 0:
            col = self._combo_column.currentData()
            self._import_with_column(self._current_path, col)

    def _import_with_column(self, path: str, column: int):
        is_csv = path.lower().endswith(".csv")
        try:
            if is_csv:
                valid, skipped, row_data = import_from_csv(path, column=column, return_rows=True)
            else:
                valid, skipped, row_data = import_from_excel(path, column=column, return_rows=True)
        except Exception as e:
            QMessageBox.critical(self, "Błąd importu", str(e))
            return

        valid, dup_count = deduplicate_numbers(valid)
        dup_text = f", {dup_count} duplikatów usuniętych" if dup_count > 0 else ""

        self._lbl_summary.setText(
            f"Załadowano: {len(valid)} {self._plural_numery(len(valid))} "
            f"({len(skipped)} pominiętych{dup_text})"
        )

        if self._headers:
            self.headers_changed.emit(self._headers)

        self.numbers_changed.emit(valid, skipped, row_data)

    def _get_max_columns(self, path: str, is_csv: bool) -> int:
        if is_csv:
            import csv as csv_mod
            with open(path, "r", encoding="utf-8") as f:
                reader = csv_mod.reader(f)
                for row in reader:
                    return len(row)
            return 1
        else:
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True)
            ws = wb.active
            max_col = ws.max_column or 1
            wb.close()
            return max_col

    @staticmethod
    def _plural_numery(n: int) -> str:
        if n == 1:
            return "numer"
        elif 2 <= n % 10 <= 4 and not (12 <= n % 100 <= 14):
            return "numery"
        return "numerów"

    def set_enabled(self, enabled: bool):
        self._btn_import.setEnabled(enabled)
        self._combo_column.setEnabled(enabled and bool(self._current_path))
