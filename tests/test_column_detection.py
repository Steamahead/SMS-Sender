import os
import tempfile

import pytest
from openpyxl import Workbook

from core.excel_importer import detect_phone_column


class TestDetectPhoneColumn:
    def _create_xlsx(self, rows: list[list]) -> str:
        wb = Workbook()
        ws = wb.active
        for r_idx, row in enumerate(rows, start=1):
            for c_idx, val in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=val)
        path = os.path.join(tempfile.mkdtemp(), "test.xlsx")
        wb.save(path)
        return path

    def test_phone_in_column_a(self):
        path = self._create_xlsx([
            ["+48512345678", "Jan", "Kowalski"],
            ["601234567", "Anna", "Nowak"],
            ["48701234567", "Piotr", "Wisniewski"],
        ])
        assert detect_phone_column(path) == 0

    def test_phone_in_column_b(self):
        path = self._create_xlsx([
            ["Jan", "+48512345678", "Firma A"],
            ["Anna", "601234567", "Firma B"],
            ["Piotr", "48701234567", "Firma C"],
        ])
        assert detect_phone_column(path) == 1

    def test_phone_in_column_c(self):
        path = self._create_xlsx([
            ["Jan", "Kowalski", "512345678"],
            ["Anna", "Nowak", "601234567"],
        ])
        assert detect_phone_column(path) == 2

    def test_no_phone_column(self):
        path = self._create_xlsx([
            ["Jan", "Kowalski", "Firma"],
            ["Anna", "Nowak", "Firma B"],
        ])
        assert detect_phone_column(path) is None

    def test_empty_file(self):
        path = self._create_xlsx([])
        assert detect_phone_column(path) is None

    def test_mixed_data_picks_best_column(self):
        path = self._create_xlsx([
            ["12345", "+48512345678", "text"],
            ["abc", "601234567", "text"],
            ["xyz", "48701234567", "text"],
        ])
        assert detect_phone_column(path) == 1


from core.excel_importer import get_column_headers, import_from_excel


class TestGetColumnHeaders:
    def _create_xlsx(self, rows: list[list]) -> str:
        wb = Workbook()
        ws = wb.active
        for r_idx, row in enumerate(rows, start=1):
            for c_idx, val in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=val)
        path = os.path.join(tempfile.mkdtemp(), "test.xlsx")
        wb.save(path)
        return path

    def test_has_headers(self):
        path = self._create_xlsx([
            ["Imie", "Telefon", "Firma"],
            ["Jan", "+48512345678", "ABC"],
        ])
        headers = get_column_headers(path)
        assert headers == ["Imie", "Telefon", "Firma"]

    def test_numeric_first_row_no_headers(self):
        path = self._create_xlsx([
            ["+48512345678", "601234567"],
            ["+48701234567", "501234567"],
        ])
        headers = get_column_headers(path)
        assert headers is None

    def test_empty_file(self):
        path = self._create_xlsx([])
        headers = get_column_headers(path)
        assert headers is None


class TestImportWithColumn:
    def _create_xlsx(self, rows: list[list]) -> str:
        wb = Workbook()
        ws = wb.active
        for r_idx, row in enumerate(rows, start=1):
            for c_idx, val in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=val)
        path = os.path.join(tempfile.mkdtemp(), "test.xlsx")
        wb.save(path)
        return path

    def test_import_from_column_b(self):
        path = self._create_xlsx([
            ["Jan", "+48512345678"],
            ["Anna", "601234567"],
        ])
        valid, skipped = import_from_excel(path, column=1)
        assert len(valid) == 2

    def test_import_default_column_a(self):
        path = self._create_xlsx([
            ["+48512345678"],
            ["601234567"],
        ])
        valid, skipped = import_from_excel(path, column=0)
        assert len(valid) == 2

    def test_import_returns_all_row_data(self):
        path = self._create_xlsx([
            ["Jan", "+48512345678", "Firma A"],
            ["Anna", "601234567", "Firma B"],
        ])
        valid, skipped, row_data = import_from_excel(path, column=1, return_rows=True)
        assert len(valid) == 2
        assert row_data[0] == ["Jan", "+48512345678", "Firma A"]
        assert row_data[1] == ["Anna", "601234567", "Firma B"]
