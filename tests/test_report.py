import os
import tempfile

import pytest
from openpyxl import load_workbook

from core.report import export_report_xlsx, export_report_csv


class TestExportReportXlsx:
    def test_creates_file(self):
        path = os.path.join(tempfile.mkdtemp(), "report.xlsx")
        recipients = [
            {"number": "+48512345678", "status": "sent", "message": "Witaj!", "time": "14:32:01", "error": ""},
            {"number": "+48601234567", "status": "error", "message": "Witaj!", "time": "14:32:05", "error": "Timeout"},
        ]
        export_report_xlsx(path, recipients)
        assert os.path.exists(path)

    def test_correct_content(self):
        path = os.path.join(tempfile.mkdtemp(), "report.xlsx")
        recipients = [
            {"number": "+48512345678", "status": "sent", "message": "Witaj!", "time": "14:32:01", "error": ""},
        ]
        export_report_xlsx(path, recipients)
        wb = load_workbook(path)
        ws = wb.active
        assert ws.cell(1, 1).value == "Numer"
        assert ws.cell(1, 2).value == "Status"
        assert ws.cell(2, 1).value == "+48512345678"
        assert ws.cell(2, 2).value == "sent"
        wb.close()

    def test_empty_recipients(self):
        path = os.path.join(tempfile.mkdtemp(), "report.xlsx")
        export_report_xlsx(path, [])
        assert os.path.exists(path)


class TestExportReportCsv:
    def test_creates_file(self):
        path = os.path.join(tempfile.mkdtemp(), "report.csv")
        recipients = [
            {"number": "+48512345678", "status": "sent", "message": "Witaj!", "time": "14:32:01", "error": ""},
        ]
        export_report_csv(path, recipients)
        assert os.path.exists(path)

    def test_correct_content(self):
        path = os.path.join(tempfile.mkdtemp(), "report.csv")
        recipients = [
            {"number": "+48512345678", "status": "sent", "message": "Witaj!", "time": "14:32:01", "error": ""},
        ]
        export_report_csv(path, recipients)
        with open(path, "r", encoding="utf-8") as f:
            lines = f.readlines()
        assert "Numer" in lines[0]
        assert "+48512345678" in lines[1]


class TestFormulaInjection:
    def test_xlsx_neutralizes_formula_in_message(self):
        path = os.path.join(tempfile.mkdtemp(), "report.xlsx")
        recipients = [
            {"number": "+48512345678", "status": "sent",
             "message": "=HYPERLINK(\"http://evil\")", "time": "10:00:00", "error": ""},
        ]
        export_report_xlsx(path, recipients)
        wb = load_workbook(path)
        ws = wb.active
        assert ws.cell(2, 3).value.startswith("'=")  # apostrophe-escaped, inert
        wb.close()

    def test_csv_neutralizes_formula_in_error(self):
        path = os.path.join(tempfile.mkdtemp(), "report.csv")
        recipients = [
            {"number": "+48512345678", "status": "error",
             "message": "Hi", "time": "10:00:00", "error": "@SUM(1+1)"},
        ]
        export_report_csv(path, recipients)
        with open(path, "r", encoding="utf-8") as f:
            content = f.read()
        assert "'@SUM(1+1)" in content

    def test_normal_message_unchanged(self):
        path = os.path.join(tempfile.mkdtemp(), "report.csv")
        recipients = [
            {"number": "+48512345678", "status": "sent",
             "message": "Dzień dobry", "time": "10:00:00", "error": ""},
        ]
        export_report_csv(path, recipients)
        with open(path, "r", encoding="utf-8") as f:
            content = f.read()
        assert "Dzień dobry" in content
        assert "'Dzień" not in content
