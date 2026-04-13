import csv
import os

import phonenumbers
from openpyxl import load_workbook


def normalize_phone_number(raw) -> str:
    """Normalize a phone number to +48XXXXXXXXX format."""
    text = str(raw).strip().replace(" ", "").replace("-", "").replace("(", "").replace(")", "").replace(".", "")

    if not text.startswith("+"):
        if text.startswith("48") and len(text) == 11:
            text = "+" + text
        elif len(text) == 9 and text.isdigit():
            text = "+48" + text
        else:
            text = "+" + text

    return text


def validate_phone_number(raw) -> tuple[bool, str | None, str | None]:
    """Validate and normalize a phone number.

    Returns: (is_valid, normalized_number_or_None, error_reason_or_None)
    """
    if raw is None or str(raw).strip() == "":
        return False, None, "Pusty numer"

    try:
        normalized = normalize_phone_number(raw)
        parsed = phonenumbers.parse(normalized, "PL")

        if not phonenumbers.is_valid_number(parsed):
            return False, None, f"Nieprawidlowy numer: {raw}"

        formatted = phonenumbers.format_number(
            parsed, phonenumbers.PhoneNumberFormat.E164
        )
        return True, formatted, None

    except phonenumbers.NumberParseException:
        return False, None, f"Nie mozna sparsowac: {raw}"


def detect_phone_column(path: str, max_rows: int = 50) -> int | None:
    """Auto-detect which column contains phone numbers.
    Scans up to max_rows rows, returns 0-based column index with most valid numbers.
    Returns None if no column has at least 2 valid numbers.
    """
    if not os.path.exists(path):
        raise FileNotFoundError(f"Plik nie istnieje: {path}")

    wb = load_workbook(path, read_only=True)
    ws = wb.active

    scores: dict[int, int] = {}

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_idx > max_rows:
            break
        for col_idx, val in enumerate(row):
            if val is None or str(val).strip() == "":
                continue
            is_valid, _, _ = validate_phone_number(val)
            if is_valid:
                scores[col_idx] = scores.get(col_idx, 0) + 1

    wb.close()

    if not scores:
        return None

    best_col = max(scores, key=scores.get)
    if scores[best_col] < 2:
        return None
    return best_col


def import_from_excel(
    path: str,
    column: int = 0,
    return_rows: bool = False,
) -> tuple[list[str], list[dict]] | tuple[list[str], list[dict], list[list]]:
    """Import phone numbers from .xlsx file.

    Args:
        path: Path to .xlsx file
        column: 0-based column index to read phone numbers from
        return_rows: If True, also return all row data (for personalization)

    Returns: (valid_numbers, skipped_entries) or (valid_numbers, skipped_entries, row_data)
    """
    if not os.path.exists(path):
        raise FileNotFoundError(f"Plik nie istnieje: {path}")

    wb = load_workbook(path, read_only=True)
    ws = wb.active

    valid = []
    skipped = []
    row_data = []

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        row_list = list(row)

        if column >= len(row_list):
            skipped.append({"row": row_idx, "value": "", "reason": "Brak kolumny"})
            continue

        raw = row_list[column]
        if raw is None or str(raw).strip() == "":
            skipped.append({"row": row_idx, "value": "", "reason": "Pusty numer"})
            continue

        is_valid, normalized, reason = validate_phone_number(raw)
        if is_valid:
            valid.append(normalized)
            if return_rows:
                row_data.append([str(v) if v is not None else "" for v in row_list])
        else:
            skipped.append({"row": row_idx, "value": str(raw), "reason": reason})

    wb.close()

    if return_rows:
        return valid, skipped, row_data
    return valid, skipped


def deduplicate_numbers(numbers: list[str]) -> tuple[list[str], int]:
    """Remove duplicate phone numbers preserving order.
    Returns: (unique_numbers, count_of_removed_duplicates)
    """
    seen = set()
    unique = []
    for num in numbers:
        if num not in seen:
            seen.add(num)
            unique.append(num)
    return unique, len(numbers) - len(unique)


def get_column_headers(path: str) -> list[str] | None:
    """Read first row of Excel file and return as headers if they look like text.
    Returns None if the first row looks like data (contains valid phone numbers).
    """
    if not os.path.exists(path):
        raise FileNotFoundError(f"Plik nie istnieje: {path}")

    wb = load_workbook(path, read_only=True)
    ws = wb.active

    first_row = []
    for row in ws.iter_rows(max_row=1, values_only=True):
        first_row = [str(v) if v is not None else "" for v in row]
        break

    wb.close()

    if not first_row:
        return None

    phone_count = sum(
        1 for val in first_row
        if val.strip() and validate_phone_number(val)[0]
    )

    if phone_count > 0:
        return None

    return first_row


def detect_phone_column_csv(path: str, max_rows: int = 50) -> int | None:
    """Auto-detect which column in a CSV contains phone numbers."""
    if not os.path.exists(path):
        raise FileNotFoundError(f"Plik nie istnieje: {path}")

    scores: dict[int, int] = {}

    with open(path, "r", encoding="utf-8") as f:
        reader = csv.reader(f)
        for row_idx, row in enumerate(reader, start=1):
            if row_idx > max_rows:
                break
            for col_idx, val in enumerate(row):
                if not val.strip():
                    continue
                is_valid, _, _ = validate_phone_number(val)
                if is_valid:
                    scores[col_idx] = scores.get(col_idx, 0) + 1

    if not scores:
        return None

    best_col = max(scores, key=scores.get)
    if scores[best_col] < 2:
        return None
    return best_col


def import_from_csv(
    path: str,
    column: int = 0,
    return_rows: bool = False,
) -> tuple[list[str], list[dict]] | tuple[list[str], list[dict], list[list]]:
    """Import phone numbers from .csv file."""
    if not os.path.exists(path):
        raise FileNotFoundError(f"Plik nie istnieje: {path}")

    valid = []
    skipped = []
    row_data = []

    with open(path, "r", encoding="utf-8") as f:
        reader = csv.reader(f)
        for row_idx, row in enumerate(reader, start=1):
            if column >= len(row) or not row[column].strip():
                skipped.append({"row": row_idx, "value": "", "reason": "Pusty numer"})
                continue

            raw = row[column].strip()
            is_valid, normalized, reason = validate_phone_number(raw)
            if is_valid:
                valid.append(normalized)
                if return_rows:
                    row_data.append(row)
            else:
                skipped.append({"row": row_idx, "value": raw, "reason": reason})

    if return_rows:
        return valid, skipped, row_data
    return valid, skipped
