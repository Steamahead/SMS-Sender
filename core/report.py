import csv
import os

from openpyxl import Workbook


_HEADERS = ["Numer", "Status", "Tresc", "Czas", "Blad"]

# Leading characters that spreadsheet apps interpret as the start of a formula.
_FORMULA_TRIGGERS = ("=", "+", "-", "@", "\t", "\r")


def _sanitize_cell(value) -> str:
    """Neutralize CSV/Excel formula injection.

    Free-text fields (message, error) may start with characters that Excel
    treats as a formula. Prefix such values with an apostrophe so they are
    rendered as literal text when the report is opened.
    """
    text = "" if value is None else str(value)
    if text and text[0] in _FORMULA_TRIGGERS:
        return "'" + text
    return text


def export_report_xlsx(path: str, recipients: list[dict]) -> None:
    """Export sending report to Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Raport SMS"

    for col_idx, header in enumerate(_HEADERS, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    for row_idx, r in enumerate(recipients, start=2):
        ws.cell(row=row_idx, column=1, value=r.get("number", ""))
        ws.cell(row=row_idx, column=2, value=r.get("status", ""))
        ws.cell(row=row_idx, column=3, value=_sanitize_cell(r.get("message", "")))
        ws.cell(row=row_idx, column=4, value=r.get("time", ""))
        ws.cell(row=row_idx, column=5, value=_sanitize_cell(r.get("error", "")))

    os.makedirs(os.path.dirname(path), exist_ok=True) if os.path.dirname(path) else None
    wb.save(path)


def export_report_csv(path: str, recipients: list[dict]) -> None:
    """Export sending report to CSV file."""
    os.makedirs(os.path.dirname(path), exist_ok=True) if os.path.dirname(path) else None

    with open(path, "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(_HEADERS)
        for r in recipients:
            writer.writerow([
                r.get("number", ""),
                r.get("status", ""),
                _sanitize_cell(r.get("message", "")),
                r.get("time", ""),
                _sanitize_cell(r.get("error", "")),
            ])
